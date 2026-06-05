#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path.cwd()
MEMORY_ROOT = ROOT / ".yeyo-memory"
CARDS_DIR = MEMORY_ROOT / "cards"
REPORTS_DIR = MEMORY_ROOT / "reports"


DOC_CODE_RE = re.compile(
    r"\b(?:\d{4}[-_][A-Z]{3}[-_][A-Z]{3}[-_][A-Z0-9]{2,4}[-_][A-Z0-9]{2,4}[-_][A-Z0-9]{3,6}|\d{3}[-_](?:HTF|MS|N)[-_][A-Z0-9]+[-_][A-Z0-9]+[-_]\d+|[A-Z]\d{2}[-_]\d{2}[-_]\d{4}[-_][A-Z]{2}[-_][A-Z]{3}[-_]\d{6})\b",
    re.I,
)

HEADER_KEYWORDS = {
    "document": ["document", "doc.", "doc ", "document no", "document number", "document nº", "document n"],
    "drawing": ["drawing", "dwg", "pln", "plan"],
    "code": ["code", "number", "no.", "nº"],
    "title": ["title", "description", "document title", "drawing title", "name"],
    "revision": ["revision", "rev", "rev."],
    "status": ["status", "purpose", "issue", "ifc", "submitted"],
    "date": ["date"],
    "discipline": ["discipline", "area", "system", "unit", "category"],
}

INDEX_NAME_HINTS = [
    "document list",
    "lista de documentos",
    "lista documentos",
    "as built list",
    "edl",
    "support list",
    "mto",
    "equipment list",
    "valve list",
    "instrument list",
    "signal list",
    "cable list",
    "packing and shipping list",
]


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def norm_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def norm_code(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def find_doc_code(values: list[Any]) -> str:
    for value in values:
        text = norm(value)
        match = DOC_CODE_RE.search(text)
        if match:
            return match.group(0).replace("_", "-")
    return ""


def classify_index_kind(path: str) -> str:
    text = norm_key(path)
    if "support list" in text or "soporte" in text:
        return "support-list"
    if "document list" in text or "lista de documentos" in text or "lista documentos" in text:
        return "document-list"
    if "as built list" in text:
        return "as-built-list"
    if "edl" in text:
        return "edl"
    if "mto" in text:
        return "mto"
    if "valve list" in text:
        return "valve-list"
    if "instrument" in text or "signal list" in text or "cable list" in text:
        return "ic-list"
    if "equipment list" in text:
        return "equipment-list"
    return "index-like-workbook"


def is_index_candidate(card: dict[str, Any]) -> bool:
    if card.get("ext") not in {"xlsx", "xlsm"}:
        return False
    text = norm_key(f"{card.get('path')} {card.get('title')}")
    return any(hint in text for hint in INDEX_NAME_HINTS) or "/lis/" in text or "lista" in text


def header_score(values: list[Any]) -> int:
    text_values = [norm_key(value) for value in values]
    score = 0
    for variants in HEADER_KEYWORDS.values():
        if any(any(variant in value for variant in variants) for value in text_values):
            score += 1
    return score


def map_headers(values: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(values):
        header = norm_key(value)
        if not header:
            continue
        for field, variants in HEADER_KEYWORDS.items():
            if field in mapping:
                continue
            if any(variant in header for variant in variants):
                mapping[field] = index
    return mapping


def find_header_row(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    best: tuple[int, int, dict[str, int]] | None = None
    for index, values in enumerate(rows[:30]):
        score = header_score(values)
        mapping = map_headers(values)
        if score >= 2 and ("title" in mapping or "document" in mapping or "drawing" in mapping):
            candidate = (score, index, mapping)
            if best is None or candidate[0] > best[0]:
                best = candidate
    if best is None:
        return None
    return best[1], best[2]


def cell(row: list[Any], mapping: dict[str, int], key: str) -> str:
    index = mapping.get(key)
    if index is None or index >= len(row):
        return ""
    return norm(row[index])


def extract_entries(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    entries: list[dict[str, Any]] = []
    sheets_info: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            header = find_header_row(rows)
            if header is None:
                sheets_info.append({"sheet": sheet.title, "rows": len(rows), "entries": 0, "header_found": False})
                continue
            header_index, mapping = header
            sheet_entries = 0
            for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                values = [norm(value) for value in row]
                if not any(values):
                    continue
                doc_code = cell(row, mapping, "document") or cell(row, mapping, "drawing") or find_doc_code(values)
                title = cell(row, mapping, "title")
                revision = cell(row, mapping, "revision")
                status = cell(row, mapping, "status")
                date = cell(row, mapping, "date")
                discipline = cell(row, mapping, "discipline")
                if not doc_code and not title:
                    continue
                if not doc_code and len(title) < 8:
                    continue
                entries.append(
                    {
                        "sheet": sheet.title,
                        "row": row_number,
                        "doc_code": doc_code.replace("_", "-"),
                        "title": title,
                        "revision": revision,
                        "status": status,
                        "date": date,
                        "discipline": discipline,
                        "raw": " | ".join(value for value in values if value)[:1200],
                    }
                )
                sheet_entries += 1
            sheets_info.append({"sheet": sheet.title, "rows": len(rows), "entries": sheet_entries, "header_found": True})
    finally:
        workbook.close()
    return entries, sheets_info


def load_cards() -> list[dict[str, Any]]:
    return [json.loads(path.read_text(encoding="utf-8")) for path in sorted(CARDS_DIR.glob("*.json"))]


def build_card_lookup(cards: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    lookup: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for card in cards:
        values = [card.get("doc_code") or "", Path(card.get("path") or "").stem]
        code = find_doc_code(values) or (card.get("doc_code") or "")
        key = norm_code(code)
        if key:
            lookup[key].append(card)
    return lookup


def match_cards(entry: dict[str, Any], lookup: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    key = norm_code(entry.get("doc_code"))
    if not key:
        return []
    matches = lookup.get(key, [])
    if matches:
        return matches
    partial = []
    for candidate_key, cards in lookup.items():
        if key and (key in candidate_key or candidate_key in key):
            partial.extend(cards)
    return partial[:10]


def infer_revision(rel_path: str) -> str | None:
    name = Path(rel_path).name
    patterns = [
        r"\bRev[ ._-]?([A-Z]?\d{1,3}|[A-Z])\b",
        r"\bR([A-Z]?\d{1,3})\b",
        r"\brev([A-Z]?\d{1,3})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, name, re.I)
        if match:
            return match.group(0)
    return None


def parse_revision_score(rev_str: str | None, path: str) -> tuple[int, int]:
    if not rev_str:
        rev_str = infer_revision(path)
    if not rev_str:
        return (-1, 0)
    rev_str = rev_str.lower()
    nums = re.findall(r"\d+", rev_str)
    if nums:
        major = int(nums[0])
        minor = int(nums[1]) if len(nums) > 1 else 0
        return (major, minor)
    alpha_match = re.search(r"\b(?:rev|r|revision)[ ._-]?([a-z])\b", rev_str)
    if alpha_match:
        return (ord(alpha_match.group(1)) - ord('a'), 0)
    return (0, 0)


def card_priority_key(card: dict[str, Any]) -> tuple[int, tuple[int, int], int]:
    path = card.get("path", "").lower()
    ext = card.get("ext", "").lower()
    rev = card.get("revision")
    
    path_score = 0
    if "as built" in path or "as-built" in path:
        path_score = 10
    elif "sello pe" in path or "pe 08" in path or "sello_pe" in path:
        path_score = 8
    elif "condado" in path or "permit" in path:
        path_score = 5
    elif any(folder in path for folder in ["01 gd ae", "02 gd loi", "03 gd fabricacion", "04 gd fabricacon", "05 gd fabricaci"]):
        path_score = 2
        
    rev_score = parse_revision_score(rev, path)
    
    ext_score = 0
    if ext in {"pdf", "xlsx"}:
        ext_score = 3
    elif ext in {"docx", "xls"}:
        ext_score = 2
    elif ext in {"zip", "7z"}:
        ext_score = 1
        
    return (path_score, rev_score, ext_score)


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def main() -> int:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    cards = load_cards()
    lookup = build_card_lookup(cards)
    index_cards = [card for card in cards if is_index_candidate(card)]

    index_rows: list[dict[str, Any]] = []
    entry_rows: list[dict[str, Any]] = []
    match_rows: list[dict[str, Any]] = []
    kind_counts = Counter()
    matched_card_ids: set[str] = set()

    for card in index_cards:
        path = ROOT / card["path"]
        try:
            entries, sheets = extract_entries(path)
            error = ""
        except Exception as exc:
            entries, sheets, error = [], [], str(exc)[:1000]

        index_kind = classify_index_kind(card["path"])
        kind_counts[index_kind] += 1
        index_rows.append(
            {
                "path": card["path"],
                "index_kind": index_kind,
                "entries": len(entries),
                "sheets": len(sheets),
                "error": error,
                "card_path": card["card_path"],
            }
        )

        for entry in entries:
            row = {
                "index_path": card["path"],
                "index_kind": index_kind,
                **entry,
            }
            entry_rows.append(row)

            matches = match_cards(entry, lookup)
            if matches:
                best_match = max(matches, key=card_priority_key)
                matches = [best_match]
            for matched in matches:
                matched_card_ids.add(matched["id"])
                match_rows.append(
                    {
                        "index_path": card["path"],
                        "index_kind": index_kind,
                        "sheet": entry["sheet"],
                        "row": entry["row"],
                        "index_doc_code": entry["doc_code"],
                        "index_title": entry["title"],
                        "index_revision": entry["revision"],
                        "index_status": entry["status"],
                        "matched_path": matched["path"],
                        "matched_ext": matched["ext"],
                        "matched_status": matched["status"],
                        "matched_card_path": matched["card_path"],
                    }
                )

                metadata = matched.setdefault("metadata", {})
                index_matches = metadata.setdefault("index_matches", [])
                fingerprint = f"{card['path']}::{entry['sheet']}::{entry['row']}"
                if not any(item.get("fingerprint") == fingerprint for item in index_matches):
                    index_matches.append(
                        {
                            "fingerprint": fingerprint,
                            "index_path": card["path"],
                            "index_kind": index_kind,
                            "sheet": entry["sheet"],
                            "row": entry["row"],
                            "doc_code": entry["doc_code"],
                            "title": entry["title"],
                            "revision": entry["revision"],
                            "status": entry["status"],
                            "date": entry["date"],
                            "discipline": entry["discipline"],
                        }
                    )
                    (ROOT / matched["card_path"]).write_text(json.dumps(matched, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    write_csv(REPORTS_DIR / "document-indexes.csv", index_rows, ["path", "index_kind", "entries", "sheets", "error", "card_path"])
    write_csv(
        REPORTS_DIR / "document-index-entries.csv",
        entry_rows,
        ["index_path", "index_kind", "sheet", "row", "doc_code", "title", "revision", "status", "date", "discipline", "raw"],
    )
    write_csv(
        REPORTS_DIR / "document-index-matches.csv",
        match_rows,
        [
            "index_path",
            "index_kind",
            "sheet",
            "row",
            "index_doc_code",
            "index_title",
            "index_revision",
            "index_status",
            "matched_path",
            "matched_ext",
            "matched_status",
            "matched_card_path",
        ],
    )

    lines = [
        "# Indices documentales XLS/XLSX",
        "",
        f"- Generado: {datetime.now(timezone.utc).isoformat()}",
        f"- Libros candidatos: {len(index_cards)}",
        f"- Filas de indice extraidas: {len(entry_rows)}",
        f"- Cruces con documentos existentes: {len(match_rows)}",
        f"- Documentos unicos enriquecidos por indice: {len(matched_card_ids)}",
        "",
        "## Tipos de indice",
        "",
        "| Tipo | Libros |",
        "|---|---:|",
        *[f"| {kind} | {count} |" for kind, count in kind_counts.most_common()],
        "",
        "Archivos generados:",
        "",
        "- `document-indexes.csv`: libros detectados como indices.",
        "- `document-index-entries.csv`: filas extraidas de indices.",
        "- `document-index-matches.csv`: cruces entre indice y documentos del repositorio.",
    ]
    (REPORTS_DIR / "document-indexes.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
